from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django_filters import rest_framework as django_filters
from django.db import transaction
from django.db.models import Q, Sum, Count, F 
from django.utils import timezone
from datetime import datetime, timedelta
import io 
from django.http import HttpResponse 
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from .models import Repuesto, MovimientoInventario, AlertaStock

from .serializers import (RepuestoSerializer,MovimientoInventarioSerializer,AlertaStockSerializer)


class RepuestoViewSet(viewsets.ModelViewSet):

    """
    ViewSet para gestión completa de repuestos
    """
    queryset = Repuesto.objects.all()
    serializer_class = RepuestoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    # ✅ ARREGLO: Remover 'necesita_reposicion' porque es una property, no un campo del modelo
    filterset_fields = ['activo', 'marca']
    
    search_fields = ['nombre', 'descripcion', 'marca', 'modelo', 'codigo_barras']
    ordering_fields = ['nombre', 'stock_actual', 'stock_minimo_seguridad', 'creado_en']
    ordering = ['nombre']

    def perform_create(self, serializer):
        """Asignar usuario creador al crear repuesto"""
        try:
            print(f"DEBUG perform_create: Usuario = {self.request.user}")
            print(f"DEBUG perform_create: Usuario ID = {self.request.user.id}")
            print(f"DEBUG perform_create: Datos del serializer = {serializer.validated_data}")
            
            # ✅ IMPORTANTE: Asignar el usuario creador
            repuesto = serializer.save(creado_por=self.request.user)
            
            print(f"DEBUG perform_create: Repuesto creado = {repuesto}")
            print(f"DEBUG perform_create: Repuesto ID = {repuesto.id}")
            
        except Exception as e:
            print(f"ERROR en perform_create: {e}")
            import traceback
            traceback.print_exc()
            raise

    def get_permissions(self):
        """
        Permisos específicos por acción:
        - Lectura: Todos los autenticados
        - Escritura: Solo Supervisor y Encargado de Bodega
        """
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            user = self.request.user
            
            # ✅ ARREGLO: Verificar permisos correctamente
            print(f"DEBUG: Usuario {user.username}, is_superuser: {user.is_superuser}")
            if hasattr(user, 'rol'):
                print(f"DEBUG: rol: {user.rol}")
            print(f"DEBUG: grupos: {[g.name for g in user.groups.all()]}")
            
            if not (user.is_superuser or 
                   (hasattr(user, 'rol') and user.rol in ['SUPER_ADMIN', 'SUPERVISOR', 'ENCARGADO_BODEGA']) or
                   user.groups.filter(name__in=['Supervisor', 'Encargado de Bodega']).exists()):
                self.permission_denied(
                    self.request, 
                    message="Solo supervisores y encargados de bodega pueden modificar repuestos."
                )
        return super().get_permissions()

    def get_queryset(self):
        """
        Override para filtros personalizados que no se pueden hacer con filterset_fields
        """
        queryset = super().get_queryset()
        
        # ✅ NUEVO: Filtro personalizado para necesita_reposicion
        necesita_reposicion = self.request.query_params.get('necesita_reposicion')
        if necesita_reposicion is not None:
            if necesita_reposicion.lower() == 'true':
                # Filtrar repuestos que necesitan reposición (stock <= stock_minimo)
                queryset = queryset.filter(stock_actual__lte=F('stock_minimo_seguridad'))
            elif necesita_reposicion.lower() == 'false':
                # Filtrar repuestos que NO necesitan reposición (stock > stock_minimo)
                queryset = queryset.filter(stock_actual__gt=F('stock_minimo_seguridad'))
        
        return queryset

    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Obtener estadísticas generales del inventario"""
        queryset = self.get_queryset()
        
        try:
            # ✅ CORREGIR: Calcular estadísticas de manera segura
            total_repuestos = queryset.count()
            repuestos_activos = queryset.filter(activo=True).count()
            
            # Calcular alertas de stock bajo
            alertas_stock_bajo = queryset.filter(
                stock_actual__lte=F('stock_minimo_seguridad')
            ).count()
            
            # Calcular valor total del inventario de forma segura
            valor_total = 0
            for repuesto in queryset.filter(activo=True, costo_unitario__isnull=False):
                valor_total += (repuesto.costo_unitario or 0) * repuesto.stock_actual
            
            repuestos_sin_stock = queryset.filter(stock_actual=0).count()
            
            stats = {
                'total_repuestos': total_repuestos,
                'repuestos_activos': repuestos_activos,
                'alertas_stock_bajo': alertas_stock_bajo,
                'valor_total_inventario': float(valor_total),
                'repuestos_sin_stock': repuestos_sin_stock,
            }
            
            print(f"DEBUG: Estadísticas calculadas: {stats}")
            return Response(stats)
            
        except Exception as e:
            print(f"ERROR en estadísticas: {e}")
            import traceback
            traceback.print_exc()
            
            return Response({
                'total_repuestos': 0,
                'repuestos_activos': 0,
                'alertas_stock_bajo': 0,
                'valor_total_inventario': 0.0,
                'repuestos_sin_stock': 0,
            }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def stock_bajo(self, request):
        """Listar repuestos con stock bajo"""
        repuestos_bajo_stock = self.get_queryset().filter(
            stock_actual__lte=F('stock_minimo_seguridad'),
            activo=True
        )
        
        serializer = self.get_serializer(repuestos_bajo_stock, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def ajustar_stock(self, request, pk=None):
        """Endpoint para ajustes manuales de stock (requiere autorización)"""
        repuesto = self.get_object()
        cantidad = request.data.get('cantidad')
        tipo_ajuste = request.data.get('tipo_ajuste')
        observaciones = request.data.get('observaciones', '')
        
        if not cantidad or not tipo_ajuste:
            return Response(
                {'error': 'Cantidad y tipo de ajuste son requeridos'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            cantidad = int(cantidad)
        except ValueError:
            return Response(
                {'error': 'La cantidad debe ser un número entero'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user = request.user
        if not (user.is_superuser or 
               (hasattr(user, 'rol') and user.rol in ['SUPER_ADMIN', 'SUPERVISOR']) or
               user.groups.filter(name='Supervisor').exists()):
            return Response(
                {'error': 'Solo supervisores pueden autorizar ajustes'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        tipo_movimiento = f'AJUSTE_{tipo_ajuste}'
        MovimientoInventario.objects.create(
            repuesto=repuesto,
            tipo_movimiento=tipo_movimiento,
            cantidad=cantidad,
            registrado_por=request.user,
            autorizado_por=request.user,
            observaciones=observaciones
        )
        
        return Response({'message': 'Ajuste realizado exitosamente'})


    def destroy(self, request, *args, **kwargs):
        """
        Eliminar repuesto con lógica especial para Super Administrador
        """
        repuesto = self.get_object()
        user = request.user
        
        # ✅ VERIFICACIÓN DE PERMISOS
        is_super_admin = (user.is_superuser or 
                         (hasattr(user, 'rol') and user.rol == 'SUPER_ADMIN'))
        is_encargado_bodega = (user.groups.filter(name='Encargado de Bodega').exists() or
                              (hasattr(user, 'rol') and user.rol == 'ENCARGADO_BODEGA'))
        
        if not (is_super_admin or is_encargado_bodega):
            return Response(
                {'error': 'Solo super administradores y encargados de bodega pueden eliminar repuestos'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # 🔥 LÓGICA ESPECIAL: Super Admin puede eliminar SIN validaciones
        if is_super_admin:
            try:
                with transaction.atomic():
                    repuesto_info = {
                        'id': repuesto.id,
                        'nombre': repuesto.nombre,
                        'marca': repuesto.marca,
                        'modelo': repuesto.modelo,
                        'stock_actual': repuesto.stock_actual,
                        'movimientos_count': MovimientoInventario.objects.filter(repuesto=repuesto).count(),
                        'alertas_count': AlertaStock.objects.filter(repuesto=repuesto).count()
                    }
                    
                    # Eliminar TODAS las alertas relacionadas
                    AlertaStock.objects.filter(repuesto=repuesto).delete()
                    
                    # Eliminar TODOS los movimientos relacionados
                    MovimientoInventario.objects.filter(repuesto=repuesto).delete()
                    
                    # Eliminar el repuesto
                    repuesto.delete()
                    
                    # Log especial para Super Admin
                    print(f"🔥 ELIMINACIÓN FORZADA POR SUPER ADMIN: {repuesto_info['nombre']} (ID: {repuesto_info['id']})")
                    print(f"   Stock eliminado: {repuesto_info['stock_actual']}")
                    print(f"   Movimientos eliminados: {repuesto_info['movimientos_count']}")
                    print(f"   Alertas eliminadas: {repuesto_info['alertas_count']}")
                    print(f"   Usuario: {user.username}")
                    
                    return Response({
                        'message': f'Repuesto "{repuesto_info["nombre"]}" eliminado FORZADAMENTE por Super Administrador',
                        'deleted_repuesto': repuesto_info,
                        'forced_deletion': True,
                        'deleted_items': {
                            'movimientos': repuesto_info['movimientos_count'],
                            'alertas': repuesto_info['alertas_count'],
                            'stock_perdido': repuesto_info['stock_actual']
                        }
                    }, status=status.HTTP_200_OK)
                    
            except Exception as e:
                print(f"❌ ERROR en eliminación forzada: {e}")
                return Response({
                    'error': f'Error interno en eliminación forzada: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # 📋 VALIDACIONES NORMALES para Encargado de Bodega
        validation_errors = []
        
        # 1. Verificar stock actual
        if repuesto.stock_actual > 0:
            validation_errors.append(f'No se puede eliminar: el repuesto tiene stock actual de {repuesto.stock_actual} {repuesto.unidad_medida}')
        
        # 2. Verificar movimientos recientes (últimos 30 días)
        fecha_limite = timezone.now() - timedelta(days=30)
        movimientos_recientes = MovimientoInventario.objects.filter(
            repuesto=repuesto,
            fecha_movimiento__gte=fecha_limite
        ).count()
        
        if movimientos_recientes > 0:
            validation_errors.append(f'No se puede eliminar: el repuesto tiene {movimientos_recientes} movimientos en los últimos 30 días')
        
        # 3. Verificar alertas pendientes
        alertas_pendientes = AlertaStock.objects.filter(
            repuesto=repuesto,
            estado__in=['PENDIENTE', 'NOTIFICADA']
        ).count()
        
        if alertas_pendientes > 0:
            validation_errors.append(f'No se puede eliminar: el repuesto tiene {alertas_pendientes} alertas pendientes')
        
        # Si hay errores de validación para Encargado de Bodega
        if validation_errors:
            return Response({
                'error': 'No se puede eliminar el repuesto',
                'validation_errors': validation_errors,
                'can_delete': False,
                'user_role': 'ENCARGADO_BODEGA'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # ✅ ELIMINAR NORMAL para Encargado de Bodega (con validaciones pasadas)
        try:
            with transaction.atomic():
                repuesto_info = {
                    'id': repuesto.id,
                    'nombre': repuesto.nombre,
                    'marca': repuesto.marca,
                    'modelo': repuesto.modelo
                }
                
                # Solo eliminar alertas resueltas
                AlertaStock.objects.filter(repuesto=repuesto, estado='RESUELTA').delete()
                
                # Eliminar el repuesto
                repuesto.delete()
                
                # Log normal
                print(f"✅ ELIMINACIÓN NORMAL: {repuesto_info['nombre']} (ID: {repuesto_info['id']}) por: {user.username}")
                
                return Response({
                    'message': f'Repuesto "{repuesto_info["nombre"]}" eliminado exitosamente',
                    'deleted_repuesto': repuesto_info,
                    'forced_deletion': False
                }, status=status.HTTP_200_OK)
                
        except Exception as e:
            print(f"❌ ERROR eliminando repuesto: {e}")
            return Response({
                'error': f'Error interno eliminando repuesto: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def validate_delete(self, request, pk=None):
        """
        Validar eliminación con lógica especial para Super Administrador
        """
        repuesto = self.get_object()
        user = request.user
        
        # Identificar tipo de usuario
        is_super_admin = (user.is_superuser or 
                         (hasattr(user, 'rol') and user.rol == 'SUPER_ADMIN'))
        is_encargado_bodega = (user.groups.filter(name='Encargado de Bodega').exists() or
                              (hasattr(user, 'rol') and user.rol == 'ENCARGADO_BODEGA'))
        
        if not (is_super_admin or is_encargado_bodega):
            return Response({
                'can_delete': False,
                'permission_error': 'Sin permisos para eliminar repuestos',
                'user_role': 'NO_PERMISSION'
            })
        
        # 🔥 SUPER ADMIN: Puede eliminar TODO siempre
        if is_super_admin:
            # Calcular información de impacto
            total_movimientos = MovimientoInventario.objects.filter(repuesto=repuesto).count()
            total_alertas = AlertaStock.objects.filter(repuesto=repuesto).count()
            alertas_pendientes = AlertaStock.objects.filter(
                repuesto=repuesto,
                estado__in=['PENDIENTE', 'NOTIFICADA']
            ).count()
            
            # Movimientos recientes
            fecha_limite = timezone.now() - timedelta(days=30)
            movimientos_recientes = MovimientoInventario.objects.filter(
                repuesto=repuesto,
                fecha_movimiento__gte=fecha_limite
            ).count()
            
            return Response({
                'can_delete': True,
                'user_role': 'SUPER_ADMIN',
                'forced_deletion': True,
                'impact_warning': {
                    'stock_actual': repuesto.stock_actual,
                    'total_movimientos': total_movimientos,
                    'movimientos_recientes': movimientos_recientes,
                    'total_alertas': total_alertas,
                    'alertas_pendientes': alertas_pendientes
                },
                'repuesto_info': {
                    'id': repuesto.id,
                    'nombre': repuesto.nombre,
                    'marca': repuesto.marca,
                    'modelo': repuesto.modelo,
                    'stock_actual': repuesto.stock_actual,
                    'unidad_medida': repuesto.unidad_medida,
                    'activo': repuesto.activo
                }
            })
        
        # 📋 ENCARGADO DE BODEGA: Validaciones normales
        validation_errors = []
        warnings = []
        
        # Stock actual
        if repuesto.stock_actual > 0:
            validation_errors.append({
                'type': 'stock_actual',
                'message': f'Tiene stock actual de {repuesto.stock_actual} {repuesto.unidad_medida}'
            })
        
        # Movimientos recientes
        fecha_limite = timezone.now() - timedelta(days=30)
        movimientos_recientes = MovimientoInventario.objects.filter(
            repuesto=repuesto,
            fecha_movimiento__gte=fecha_limite
        ).count()
        
        if movimientos_recientes > 0:
            validation_errors.append({
                'type': 'movimientos_recientes',
                'message': f'Tiene {movimientos_recientes} movimientos en los últimos 30 días'
            })
        
        # Total de movimientos (para información)
        total_movimientos = MovimientoInventario.objects.filter(repuesto=repuesto).count()
        if total_movimientos > 0:
            warnings.append({
                'type': 'historial_movimientos',
                'message': f'Se eliminarán {total_movimientos} movimientos históricos'
            })
        
        # Alertas pendientes
        alertas_pendientes = AlertaStock.objects.filter(
            repuesto=repuesto,
            estado__in=['PENDIENTE', 'NOTIFICADA']
        ).count()
        
        if alertas_pendientes > 0:
            validation_errors.append({
                'type': 'alertas_pendientes',
                'message': f'Tiene {alertas_pendientes} alertas pendientes'
            })
        
        # Alertas históricas
        alertas_historicas = AlertaStock.objects.filter(repuesto=repuesto).count()
        if alertas_historicas > 0:
            warnings.append({
                'type': 'historial_alertas',
                'message': f'Se eliminarán {alertas_historicas} alertas históricas'
            })
        
        can_delete = len(validation_errors) == 0
        
        return Response({
            'can_delete': can_delete,
            'user_role': 'ENCARGADO_BODEGA',
            'forced_deletion': False,
            'validation_errors': validation_errors,
            'warnings': warnings,
            'repuesto_info': {
                'id': repuesto.id,
                'nombre': repuesto.nombre,
                'marca': repuesto.marca,
                'modelo': repuesto.modelo,
                'stock_actual': repuesto.stock_actual,
                'unidad_medida': repuesto.unidad_medida,
                'activo': repuesto.activo
            }
        })

class MovimientoInventarioFilter(django_filters.FilterSet):
    """
    Filtros personalizados para MovimientoInventario con búsqueda avanzada
    """
    # Búsqueda por texto en múltiples campos
    search = django_filters.CharFilter(method='filter_search', label='Búsqueda general')
    
    # Filtros por fecha
    fecha_desde = django_filters.DateFilter(field_name='fecha_movimiento', lookup_expr='gte', label='Fecha desde')
    fecha_hasta = django_filters.DateFilter(field_name='fecha_movimiento', lookup_expr='lte', label='Fecha hasta')
    
    # Filtros por relaciones
    repuesto_nombre = django_filters.CharFilter(field_name='repuesto__nombre', lookup_expr='icontains', label='Nombre del repuesto')
    registrado_por_username = django_filters.CharFilter(field_name='registrado_por__username', lookup_expr='iexact', label='Usuario que registró')
    
    # Filtros por campos específicos
    proveedor = django_filters.CharFilter(lookup_expr='icontains', label='Proveedor')
    numero_factura = django_filters.CharFilter(lookup_expr='icontains', label='Número de factura')
    numero_ot = django_filters.CharFilter(lookup_expr='icontains', label='Número de OT')

    class Meta:
        model = MovimientoInventario
        fields = [
            'tipo_movimiento', 'repuesto', 'registrado_por', 'autorizado_por',
            'search', 'fecha_desde', 'fecha_hasta', 'repuesto_nombre',
            'registrado_por_username', 'proveedor', 'numero_factura', 'numero_ot'
        ]

    def filter_search(self, queryset, name, value):
        """
        Búsqueda general en múltiples campos
        """
        if not value:
            return queryset
        
        # Buscar en múltiples campos usando OR
        return queryset.filter(
            Q(repuesto__nombre__icontains=value) |
            Q(repuesto__marca__icontains=value) |
            Q(repuesto__modelo__icontains=value) |
            Q(repuesto__codigo_barras__icontains=value) |
            Q(observaciones__icontains=value) |
            Q(proveedor__icontains=value) |
            Q(numero_factura__icontains=value) |
            Q(numero_ot__icontains=value) |
            Q(registrado_por__username__icontains=value) |
            Q(registrado_por__first_name__icontains=value) |
            Q(registrado_por__last_name__icontains=value) |
            Q(autorizado_por__username__icontains=value)
        ).distinct()
    
class MovimientoInventarioViewSet(viewsets.ModelViewSet):
    """
    ViewSet mejorado para gestión de movimientos de inventario con búsqueda avanzada
    """
    queryset = MovimientoInventario.objects.select_related(
        'repuesto', 'registrado_por', 'autorizado_por'
    ).all()
    serializer_class = MovimientoInventarioSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = MovimientoInventarioFilter
    ordering_fields = ['fecha_movimiento', 'cantidad', 'stock_anterior', 'stock_posterior']
    ordering = ['-fecha_movimiento']  # Por defecto, más recientes primero

    def get_permissions(self):
        """
        Verificar permisos específicos para movimientos de inventario
        Solo Super Admin, Supervisor y Encargado de Bodega pueden acceder
        """
        user = self.request.user
        
        # Verificar diferentes formas de permisos
        is_super_user = user.is_superuser
        is_super_admin = hasattr(user, 'rol') and user.rol == 'SUPER_ADMIN'
        is_supervisor = hasattr(user, 'rol') and user.rol == 'SUPERVISOR'
        is_encargado_bodega = hasattr(user, 'rol') and user.rol == 'ENCARGADO_BODEGA'
        
        # Verificar también grupos
        has_inventory_group = user.groups.filter(
            name__in=['Supervisor', 'Encargado de Bodega']
        ).exists()
        
        has_permission = (is_super_user or is_super_admin or is_supervisor or 
                         is_encargado_bodega or has_inventory_group)
        
        if not has_permission:
            self.permission_denied(
                self.request, 
                message="Solo supervisores y encargados de bodega pueden acceder a los movimientos de inventario."
            )
        
        return super().get_permissions()

    def get_queryset(self):
        """
        Optimizar queryset con select_related para mejor rendimiento
        """
        return self.queryset.select_related(
            'repuesto', 'registrado_por', 'autorizado_por'
        )

    def perform_create(self, serializer):
        """Asignar usuario registrador al crear movimiento"""
        print(f"📝 Creando movimiento por usuario: {self.request.user.username}")
        serializer.save(registrado_por=self.request.user)

    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """
        Obtener estadísticas de movimientos - VERSIÓN CORREGIDA
        """
        try:
            queryset = self.filter_queryset(self.get_queryset())
            
            # Estadísticas generales
            total_movimientos = queryset.count()
            
            # Estadísticas por tipo - CORREGIDO
            stats_por_tipo = {}
            for tipo_code, tipo_display in MovimientoInventario.TIPOS_MOVIMIENTO:
                count = queryset.filter(tipo_movimiento=tipo_code).count()
                stats_por_tipo[tipo_code] = {
                    'nombre': tipo_display,
                    'cantidad': count
                }
            
            # Movimientos recientes (últimos 7 días)
            fecha_limite = timezone.now() - timedelta(days=7)
            movimientos_recientes = queryset.filter(
                fecha_movimiento__gte=fecha_limite
            ).count()
            
            # Usuarios más activos - CORREGIDO con Count importado
            usuarios_activos = queryset.values(
                'registrado_por__username',
                'registrado_por__first_name',
                'registrado_por__last_name'
            ).annotate(
                total_movimientos=Count('id')  # ✅ Ahora Count está importado
            ).order_by('-total_movimientos')[:5]
            
            return Response({
                'total_movimientos': total_movimientos,
                'movimientos_recientes': movimientos_recientes,
                'estadisticas_por_tipo': stats_por_tipo,
                'usuarios_mas_activos': list(usuarios_activos),
                'periodo_consulta': {
                    'desde': request.query_params.get('fecha_desde'),
                    'hasta': request.query_params.get('fecha_hasta')
                }
            })
            
        except Exception as e:
            print(f"❌ Error obteniendo estadísticas de movimientos: {e}")
            return Response({
                'error': 'Error obteniendo estadísticas',
                'detail': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    @action(detail=False, methods=['get']) 
    def exportar_excel(self, request):
        """
        Exportar movimientos en formato Excel profesional
        """
        try:
            print(f"📊 Iniciando exportación Excel...")
            print(f"👤 Usuario: {request.user.username}")
            
            # Aplicar filtros
            queryset = self.filter_queryset(self.get_queryset())
            
            # Limitar a 15000 registros para Excel (más capacidad que CSV)
            total_records = queryset.count()
            print(f"📈 Total de registros encontrados: {total_records}")
            
            if total_records > 15000:
                return Response({
                    'error': 'Demasiados registros para exportar',
                    'detail': f'Se encontraron {total_records} registros. Máximo permitido: 15,000. Use filtros para reducir los resultados.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if total_records == 0:
                return Response({
                    'error': 'No hay datos para exportar',
                    'detail': 'No se encontraron movimientos con los filtros aplicados.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Generar Excel
            return self._generate_professional_excel(queryset, total_records, request)
                
        except Exception as e:
            print(f"❌ Error exportando Excel: {e}")
            import traceback
            traceback.print_exc()
            
            error_response = HttpResponse(content_type='text/plain')
            error_response.write(f"Error generando Excel: {str(e)}")
            error_response.status_code = 500
            return error_response

    def _generate_professional_excel(self, queryset, total_records, request):
        """
        Generar archivo Excel profesional con formato avanzado
        """
        print("📊 Generando Excel profesional...")
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"
        
        # ===== ESTILOS PROFESIONALES =====
        
        # Colores corporativos
        AZUL_PRINCIPAL = '1E4A72'
        AZUL_SECUNDARIO = '2E5C8A' 
        VERDE_ENTRADA = '2E7D32'
        ROJO_SALIDA = 'C62828'
        GRIS_CLARO = 'F5F5F5'
        BLANCO = 'FFFFFF'
        
        # Fuentes
        title_font = Font(name='Calibri', size=16, bold=True, color=BLANCO)
        subtitle_font = Font(name='Calibri', size=12, bold=True, color=AZUL_PRINCIPAL)
        header_font = Font(name='Calibri', size=11, bold=True, color=BLANCO)
        data_font = Font(name='Calibri', size=10)
        small_font = Font(name='Calibri', size=9, color='666666')
        
        # Rellenos
        title_fill = PatternFill(start_color=AZUL_PRINCIPAL, end_color=AZUL_PRINCIPAL, fill_type='solid')
        header_fill = PatternFill(start_color=AZUL_SECUNDARIO, end_color=AZUL_SECUNDARIO, fill_type='solid')
        entrada_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
        salida_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
        
        # Alineaciones
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        thick_border = Border(
            left=Side(style='medium', color=AZUL_PRINCIPAL),
            right=Side(style='medium', color=AZUL_PRINCIPAL),
            top=Side(style='medium', color=AZUL_PRINCIPAL),
            bottom=Side(style='medium', color=AZUL_PRINCIPAL)
        )
        
        # ===== ENCABEZADO CORPORATIVO =====
        
        # Título principal
        ws.merge_cells('A1:M3')
        title_cell = ws['A1']
        title_cell.value = "REPORTE DE MOVIMIENTOS DE INVENTARIO"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_align
        title_cell.border = thick_border
        
        # Información del reporte
        info_row = 5
        info_data = [
            ('Fecha de generación:', timezone.now().strftime('%d/%m/%Y %H:%M:%S')),
            ('Generado por:', f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username),
            ('Total de registros:', f"{total_records:,}"),
            ('Filtros aplicados:', self._get_applied_filters_text(request))
        ]
        
        for i, (label, value) in enumerate(info_data):
            label_cell = ws.cell(row=info_row + i, column=1)
            value_cell = ws.cell(row=info_row + i, column=3)
            
            label_cell.value = label
            label_cell.font = subtitle_font
            value_cell.value = value
            value_cell.font = data_font
        
        # ===== HEADERS DE DATOS =====
        
        headers = [
            ('Fecha y Hora', 18),
            ('Tipo de Movimiento', 22), 
            ('Repuesto', 35),
            ('Marca', 15),
            ('Modelo', 15),
            ('Código', 12),
            ('Cantidad', 12),
            ('Stock Anterior', 14),
            ('Stock Actual', 14),
            ('Registrado por', 18),
            ('Proveedor', 20),
            ('N° Factura', 15),
            ('N° OT', 10),
            ('Observaciones', 30)
        ]
        
        header_row = 11
        for col, (header_text, width) in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
            # Ajustar ancho de columna
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = width
        
        # ===== DATOS CON FORMATO CONDICIONAL =====
        
        print(f"📝 Escribiendo {total_records} registros...")
        
        data_start_row = header_row + 1
        current_row = data_start_row
        
        for mov in queryset.select_related('repuesto', 'registrado_por', 'autorizado_por'):
            try:
                # Determinar tipo de movimiento para colores
                tipo_mov = mov.tipo_movimiento
                es_entrada = tipo_mov in ['ENTRADA', 'AJUSTE_POSITIVO', 'DEVOLUCION']
                es_salida = tipo_mov in ['SALIDA_USO', 'SALIDA_SOLICITUD', 'AJUSTE_NEGATIVO', 'BAJA_POR_DANHO']
                
                # Datos de la fila
                row_data = [
                    mov.fecha_movimiento.strftime('%d/%m/%Y %H:%M') if mov.fecha_movimiento else '',
                    mov.get_tipo_movimiento_display() if hasattr(mov, 'get_tipo_movimiento_display') else mov.tipo_movimiento,
                    mov.repuesto.nombre if mov.repuesto else 'N/A',
                    mov.repuesto.marca if mov.repuesto else '',
                    mov.repuesto.modelo if mov.repuesto else '',
                    mov.repuesto.codigo_barras if mov.repuesto else '',
                    mov.cantidad if mov.cantidad is not None else 0,
                    mov.stock_anterior if mov.stock_anterior is not None else 0,
                    mov.stock_posterior if mov.stock_posterior is not None else 0,
                    f"{mov.registrado_por.first_name} {mov.registrado_por.last_name}".strip() or mov.registrado_por.username if mov.registrado_por else 'N/A',
                    mov.proveedor or '',
                    mov.numero_factura or '',
                    mov.numero_ot or '',
                    mov.observaciones or ''
                ]
                
                # Escribir datos con formato
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = value
                    cell.font = data_font
                    cell.border = thin_border
                    
                    # Alineación específica por columna
                    if col in [1, 10]:  # Fecha y Usuario
                        cell.alignment = center_align
                    elif col in [7, 8, 9]:  # Cantidades y stocks
                        cell.alignment = right_align
                    else:
                        cell.alignment = left_align
                    
                    # ===== FORMATO CONDICIONAL =====
                    
                    # Colorear tipo de movimiento
                    if col == 2:
                        if es_entrada:
                            cell.fill = entrada_fill
                            cell.font = Font(name='Calibri', size=10, color=VERDE_ENTRADA, bold=True)
                        elif es_salida:
                            cell.fill = salida_fill
                            cell.font = Font(name='Calibri', size=10, color=ROJO_SALIDA, bold=True)
                    
                    # Formatear cantidad con signo y color
                    elif col == 7:
                        if isinstance(value, (int, float)) and value != 0:
                            if es_entrada:
                                cell.value = f"+{value}"
                                cell.font = Font(name='Calibri', size=10, color=VERDE_ENTRADA, bold=True)
                            elif es_salida:
                                cell.value = f"-{abs(value)}"
                                cell.font = Font(name='Calibri', size=10, color=ROJO_SALIDA, bold=True)
                    
                    # Resaltar stocks bajos (menos de 5)
                    elif col == 9:  # Stock actual
                        if isinstance(value, (int, float)) and value <= 5:
                            cell.fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')
                            cell.font = Font(name='Calibri', size=10, color='E65100', bold=True)
                
                current_row += 1
                
                # Mostrar progreso cada 1000 registros
                if (current_row - data_start_row) % 1000 == 0:
                    print(f"📊 Procesados {current_row - data_start_row} de {total_records} registros...")
                
            except Exception as row_error:
                print(f"❌ Error procesando movimiento ID {mov.id}: {row_error}")
                continue
        
        # ===== AJUSTES FINALES =====
        
        # Congelar paneles para navegación
        ws.freeze_panes = f'A{data_start_row}'
        
        # Filtros automáticos
        ws.auto_filter.ref = f'A{header_row}:N{current_row-1}'
        
        # ===== HOJA DE RESUMEN =====
        
        summary_ws = wb.create_sheet("Resumen Ejecutivo")
        
        # Título del resumen
        summary_ws.merge_cells('A1:F2')
        summary_title = summary_ws['A1']
        summary_title.value = "RESUMEN EJECUTIVO DE MOVIMIENTOS"
        summary_title.font = title_font
        summary_title.fill = title_fill
        summary_title.alignment = center_align
        summary_title.border = thick_border
        
        # ===== ESTADÍSTICAS POR TIPO =====
        
        summary_ws['A4'] = "ESTADÍSTICAS POR TIPO DE MOVIMIENTO"
        summary_ws['A4'].font = subtitle_font
        
        # Headers del resumen
        headers_resumen = ['Tipo de Movimiento', 'Cantidad', 'Porcentaje']
        for col, header in enumerate(headers_resumen, 1):
            cell = summary_ws.cell(row=6, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Calcular estadísticas
        tipos_stats = {}
        for mov in queryset:
            tipo = mov.get_tipo_movimiento_display() if hasattr(mov, 'get_tipo_movimiento_display') else mov.tipo_movimiento
            tipos_stats[tipo] = tipos_stats.get(tipo, 0) + 1
        
        # Escribir estadísticas
        row = 7
        for tipo, cantidad in sorted(tipos_stats.items(), key=lambda x: x[1], reverse=True):
            porcentaje = (cantidad / total_records) * 100 if total_records > 0 else 0
            
            summary_ws[f'A{row}'] = tipo
            summary_ws[f'B{row}'] = cantidad
            summary_ws[f'C{row}'] = f"{porcentaje:.1f}%"
            
            # Aplicar formato
            for col in ['A', 'B', 'C']:
                cell = summary_ws[f'{col}{row}']
                cell.border = thin_border
                cell.font = data_font
                if col == 'B':
                    cell.alignment = right_align
                elif col == 'C':
                    cell.alignment = center_align
            
            row += 1
        
        # Ajustar columnas del resumen
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 12
        summary_ws.column_dimensions['C'].width = 12
        
        # ===== INFORMACIÓN ADICIONAL =====
        
        summary_ws[f'A{row+2}'] = "INFORMACIÓN ADICIONAL"
        summary_ws[f'A{row+2}'].font = subtitle_font
        
        adicional_row = row + 4
        info_adicional = [
            ('Total de movimientos:', f"{total_records:,}"),
            ('Período:', self._get_period_text(request)),
            ('Usuarios involucrados:', queryset.values('registrado_por__username').distinct().count()),
            ('Repuestos afectados:', queryset.values('repuesto').distinct().count())
        ]
        
        for i, (label, value) in enumerate(info_adicional):
            summary_ws[f'A{adicional_row + i}'] = label
            summary_ws[f'B{adicional_row + i}'] = value
            summary_ws[f'A{adicional_row + i}'].font = Font(bold=True)
            summary_ws[f'B{adicional_row + i}'].font = data_font
        
        # ===== GUARDAR Y RETORNAR =====
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Crear respuesta HTTP
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"movimientos_inventario_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(output.getvalue())
        
        print(f"✅ Excel generado exitosamente: {filename}")
        print(f"📊 Registros procesados: {current_row - data_start_row}")
        print(f"📁 Tamaño del archivo: {len(output.getvalue()):,} bytes")
        
        return response

    def _get_applied_filters_text(self, request):
        """Generar texto descriptivo de filtros aplicados"""
        filtros = []
        
        if request.query_params.get('search'):
            filtros.append(f"Búsqueda: '{request.query_params.get('search')}'")
        
        if request.query_params.get('tipo_movimiento'):
            filtros.append(f"Tipo: {request.query_params.get('tipo_movimiento')}")
        
        if request.query_params.get('fecha_desde'):
            filtros.append(f"Desde: {request.query_params.get('fecha_desde')}")
        
        if request.query_params.get('fecha_hasta'):
            filtros.append(f"Hasta: {request.query_params.get('fecha_hasta')}")
        
        return '; '.join(filtros) if filtros else 'Ninguno'

    def _get_period_text(self, request):
        """Generar texto del período consultado"""
        desde = request.query_params.get('fecha_desde')
        hasta = request.query_params.get('fecha_hasta')
        
        if desde and hasta:
            return f"{desde} al {hasta}"
        elif desde:
            return f"Desde {desde}"
        elif hasta:
            return f"Hasta {hasta}"
        else:
            return "Todo el historial"
        
    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def entrada_repuestos(self, request):
        """Endpoint específico para registrar entrada de repuestos (MEJORADO)"""
        try:
            print(f"=== DEBUG ENTRADA REPUESTOS MEJORADO ===")
            print(f"Usuario: {request.user.username}")
            print(f"Datos recibidos: {request.data}")
            
            # Verificar permisos específicos para entradas
            user = request.user
            can_create_entrada = (
                user.is_superuser or
                (hasattr(user, 'rol') and user.rol in ['SUPER_ADMIN', 'SUPERVISOR', 'ENCARGADO_BODEGA']) or
                user.groups.filter(name__in=['Supervisor', 'Encargado de Bodega']).exists()
            )
            
            if not can_create_entrada:
                return Response({
                    'error': 'No tienes permisos para registrar entradas de repuestos'
                }, status=status.HTTP_403_FORBIDDEN)
        
            repuesto_id = request.data.get('repuesto_id')
            cantidad = request.data.get('cantidad')
            proveedor = request.data.get('proveedor', '')
            numero_factura = request.data.get('numero_factura', '')
            costo_unitario = request.data.get('costo_unitario')
            observaciones = request.data.get('observaciones', '')

            # Validaciones básicas
            if not repuesto_id:
                return Response({
                    'error': 'repuesto_id es requerido'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if not cantidad:
                return Response({
                    'error': 'cantidad es requerida'
                }, status=status.HTTP_400_BAD_REQUEST)

            try:
                repuesto = Repuesto.objects.get(id=repuesto_id)
                print(f"✅ Repuesto encontrado: {repuesto.nombre}")
            except Repuesto.DoesNotExist:
                return Response({
                    'error': 'Repuesto no encontrado'
                }, status=status.HTTP_404_NOT_FOUND)

            try:
                cantidad = int(cantidad)
                if cantidad <= 0:
                    raise ValueError("Cantidad debe ser positiva")
                print(f"✅ Cantidad procesada: {cantidad}")
            except ValueError as e:
                return Response({
                    'error': f'Cantidad inválida: {str(e)}'
                }, status=status.HTTP_400_BAD_REQUEST)
        
            # Crear movimiento de entrada en transacción
            with transaction.atomic():
                stock_anterior = repuesto.stock_actual
                print(f"📊 Stock anterior: {stock_anterior}")
                
                movimiento = MovimientoInventario.objects.create(
                    repuesto=repuesto,
                    tipo_movimiento='ENTRADA',
                    cantidad=cantidad,
                    registrado_por=request.user,
                    proveedor=proveedor,
                    numero_factura=numero_factura,
                    costo_unitario=costo_unitario if costo_unitario else None,
                    observaciones=observaciones
                )
                
                print(f"✅ Movimiento creado: ID {movimiento.id}")
                print(f"📊 Stock después: {repuesto.stock_actual}")
            
            # Serializar y devolver
            serializer = self.get_serializer(movimiento)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            print(f"❌ ERROR en entrada_repuestos: {e}")
            import traceback
            traceback.print_exc()
            
            return Response({
                'error': f'Error interno del servidor: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def resumen_usuario(self, request):
        """
        Obtener resumen de movimientos por usuario actual
        """
        try:
            user = request.user
            
            # Movimientos del usuario en diferentes períodos
            hoy = timezone.now().date()
            hace_7_dias = hoy - timedelta(days=7)
            hace_30_dias = hoy - timedelta(days=30)
            
            queryset = self.get_queryset().filter(registrado_por=user)
            
            resumen = {
                'usuario': {
                    'username': user.username,
                    'nombre_completo': f"{user.first_name} {user.last_name}".strip() or user.username
                },
                'movimientos_hoy': queryset.filter(fecha_movimiento__date=hoy).count(),
                'movimientos_7_dias': queryset.filter(fecha_movimiento__date__gte=hace_7_dias).count(),
                'movimientos_30_dias': queryset.filter(fecha_movimiento__date__gte=hace_30_dias).count(),
                'total_movimientos': queryset.count(),
                'tipos_mas_frecuentes': list(
                    queryset.values('tipo_movimiento').annotate(
                        total=models.Count('id')
                    ).order_by('-total')[:3]
                ),
                'ultimo_movimiento': None
            }
            
            # Último movimiento
            ultimo = queryset.first()
            if ultimo:
                resumen['ultimo_movimiento'] = {
                    'fecha': ultimo.fecha_movimiento,
                    'tipo': ultimo.get_tipo_movimiento_display(),
                    'repuesto': ultimo.repuesto.nombre,
                    'cantidad': ultimo.cantidad
                }
            
            return Response(resumen)
            
        except Exception as e:
            print(f"❌ Error obteniendo resumen de usuario: {e}")
            return Response({
                'error': 'Error obteniendo resumen'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                    
class AlertaStockViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet para consulta de alertas de stock (solo lectura)
    """
    queryset = AlertaStock.objects.all()
    serializer_class = AlertaStockSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['estado', 'repuesto']
    ordering = ['-fecha_creacion']

    @action(detail=True, methods=['post'])
    def marcar_resuelta(self, request, pk=None):
        """Marcar una alerta como resuelta"""
        alerta = self.get_object()
        observaciones = request.data.get('observaciones', '')
        
        alerta.estado = 'RESUELTA'
        alerta.fecha_resolucion = timezone.now()
        alerta.resuelta_por = request.user
        alerta.observaciones = observaciones
        alerta.save()
        
        return Response({'message': 'Alerta marcada como resuelta'})